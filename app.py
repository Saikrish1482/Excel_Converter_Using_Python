from ctypes import cdll
import psycopg2
import json
from flask import Flask, send_file, render_template_string, request, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd

app = Flask(__name__)

# --- JSON Flattener ---
def flatten_json(data, parent_key='', sep='.'):
    """
    Recursively flattens a nested JSON/dict/list into a flat dictionary.
    Example:
      {"a": {"b": [1, 2]}} -> {"a.b[0]": 1, "a.b[1]": 2}
    """
    items = []
    if isinstance(data, dict):
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            items.extend(flatten_json(v, new_key, sep=sep).items())
    elif isinstance(data, list):
        for i, v in enumerate(data):
            new_key = f"{parent_key}[{i}]"
            items.extend(flatten_json(v, new_key, sep=sep).items())
    else:
        items.append((parent_key, data))
    return dict(items)


# --- Fetch Data from DB ---
def fetch_data_from_db(state_id=3):
    conn = psycopg2.connect(
        dbname="YouDB",
        user="",
        password="",
        host="0.0.0.0",
        port="0000"
    )
    cur = conn.cursor()

    # Select multiple JSON columns
    cur.execute ("Your SQL Query here")
    
    rows = cur.fetchall()
    colnames = [desc[0] for desc in cur.description]  # get column names

    cur.close()
    conn.close()

    # Convert to list of dicts {colname: dict, ...}
    records = []
    for row in rows:
        rec = {}
        for col, val in zip(colnames, row):
            if val:
                if isinstance(val, str):
                    rec[col] = json.loads(val)
                else:
                    rec[col] = val
        records.append(rec)

    return records


# --- DB Test Route ---
@app.route('/db-test', methods=['GET'])
def db_test():
    try:
        conn = psycopg2.connect(
            dbname="YouDB",
            user="",
            password="",
            host="0.0.0.0",
            port="0000"
        )
        cur = conn.cursor()
        cur.execute("SELECT (*) FROM Table;")
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return f"DB connected successfully! Number of rows in Table: {count}"
    except Exception as e:
        return f"Error connecting to DB: {e}"


# --- Default Route ---
@app.route('/')
def home():
    return redirect(url_for('show_download_page'))


@app.route("/generate_excel", methods=["GET"])
def show_download_page():
    html = """
    <html>
        <head><title>Download Excel</title></head>
        <body style="font-family: Arial; text-align: center; margin-top: 50px;">
            <h2>Survey Export</h2>
            <p>Click below to download the Excel file:</p>
            <a href="/download_excel">
                <button style="padding: 10px 20px; font-size: 16px;">Download Overall Excel</button>
            </a>
        </body>
    </html>
    """
    return render_template_string(html)


# --- Download Excel ---
@app.route("/download_excel", methods=["GET"])
def download_excel():
    data = fetch_data_from_db()

    if not data:
        return "No data found in DB", 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    col_offset = 1
    block_names = list(set().union(*(record.keys() for record in data)))

    for block_name in block_names:
        # Add block heading (bold + bigger font)
        cell = ws.cell(row=1, column=col_offset)
        cell.value = block_name.upper()
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Collect keys for this block (flattened JSON keys)
        all_keys = set()
        for record in data:
            if block_name in record and record[block_name]:
                flat = flatten_json(record[block_name])
                all_keys.update(flat.keys())

        # Exclude unwanted keys
        all_keys = [k for k in all_keys if not (
            k.lower().endswith("_img") or 
            k.lower() in ["_img", "_img", "app_data", ""]
        )]

        # Write sub-headers (keys)
        for i, key in enumerate(all_keys):
            cell = ws.cell(row=2, column=col_offset + i)
            cell.value = key
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Write values
        row_num = 3
        for record in data:
            if block_name not in record or not record[block_name]:
                continue

            flat = flatten_json(record[block_name])
            for col_idx, key in enumerate(all_keys, start=col_offset):
                value = flat.get(key, "")

                # Convert unsupported types
                if isinstance(value, list):
                    value = ", ".join(map(str, value))
                elif isinstance(value, dict):
                    value = json.dumps(value)

                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row_num += 1

        # Leave one blank column before the next block
        col_offset += len(all_keys) + 2

    # --- Autosize all columns ---
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter  # e.g. 'A', 'B'
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    file_path = "subject.xlsx"
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=0000, debug=True)
